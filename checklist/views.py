from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.views.decorators.http import require_POST
from django.db import transaction, models
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta, date
from django.contrib.auth import login, logout
from django.contrib.auth.forms import AuthenticationForm
import json
import hashlib
import os
import uuid
from django.contrib.auth.models import User
from django.urls import reverse
from .models import (
    TodoList, Task, GForm, GQuestion, GOption, GResponse, GAnswer, GSelectedOption,
    BankQuestion, BankOption, FormPermission, FormShareLink
)
from .forms import (
    UserRegistrationForm, TodoListForm, TaskForm, GFormForm, GQuestionForm, GOptionForm, 
    FormPermissionForm, FormShareLinkForm
)
from openpyxl import Workbook

def home(request):
    return render(request, 'home.html')

def favicon_view(request):
    return HttpResponse(status=204)

@login_required
def dashboard(request):
    """Vista optimizada del dashboard con consultas eficientes"""
    from django.db.models import Count, Q, Sum
    
    # Usar select_related y prefetch_related con anotaciones para reducir consultas
    recent_todo_lists = TodoList.objects.filter(user=request.user) \
        .select_related('user') \
        .prefetch_related('tasks') \
        .annotate(
            total_tasks=Count('tasks'),
            todo_count=Count('tasks', filter=Q(tasks__status='todo')),
            progress_count=Count('tasks', filter=Q(tasks__status='progress')),
            done_count=Count('tasks', filter=Q(tasks__status='done'))
        ) \
        .order_by('-created_at')[:5]
    
    # Optimizar consulta de formularios
    forms = GForm.objects.filter(user=request.user) \
        .select_related('user') \
        .prefetch_related('questions', 'responses') \
        .annotate(
            questions_count=Count('questions', distinct=True),
            responses_count=Count('responses', distinct=True)
        ) \
        .order_by('-updated_at')[:5]
    
    # Calcular estadísticas generales de manera eficiente
    todo_lists_count = TodoList.objects.filter(user=request.user).count()
    
    completed_tasks_count = Task.objects.filter(
        todo_list__user=request.user, 
        status='done'
    ).count()
    
    forms_count = GForm.objects.filter(user=request.user).count()
    
    # Calcular respuestas en una sola consulta
    total_responses = GResponse.objects.filter(
        form__user=request.user
    ).count()
    
    return render(request, 'dashboard.html', {
        'recent_todo_lists': recent_todo_lists,
        'forms': forms,
        'todo_lists_count': todo_lists_count,
        'completed_tasks_count': completed_tasks_count,
        'forms_count': forms_count,
        'total_responses': total_responses
    })

@login_required
def todo_list_stats(request, list_id):
    """Endpoint optimizado para obtener solo las estadísticas de una lista"""
    try:
        todo_list = get_object_or_404(TodoList, id=list_id, user=request.user)
        
        # Obtener conteos en una sola consulta eficiente
        from django.db.models import Count, Q
        stats = todo_list.tasks.aggregate(
            total=Count('id'),
            todo=Count('id', filter=Q(status='todo')),
            progress=Count('id', filter=Q(status='progress')),
            done=Count('id', filter=Q(status='done'))
        )
        
        # Devolver JSON
        return JsonResponse({
            'total': stats['total'],
            'todo': stats['todo'],
            'progress': stats['progress'],
            'done': stats['done']
        })
    except TodoList.DoesNotExist:
        return JsonResponse({'error': 'Lista no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def register_view(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            login(request, form.save())
            return redirect('dashboard')
    else:
        form = UserRegistrationForm()
    return render(request, 'register.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            login(request, form.get_user())
            # Redirigir a la URL guardada en la sesión si existe
            next_url = request.session.get('next', 'dashboard')
            if 'next' in request.session:
                del request.session['next']
            return redirect(next_url)
    else:
        # Guardar la URL de redirección si viene en la solicitud
        if 'next' in request.GET:
            request.session['next'] = request.GET.get('next')
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})

def logout_view(request):
    if request.method == 'POST':
        logout(request)
        return redirect('home')
    return redirect('dashboard')

@login_required
def todo_lists(request):
    """Muestra todas las listas de tareas del usuario"""
    lists = TodoList.objects.filter(user=request.user)

    for todo_list in lists:
        todo_list.todo_count = todo_list.tasks.filter(status='todo').count()
        todo_list.progress_count = todo_list.tasks.filter(status='progress').count()
        todo_list.done_count = todo_list.tasks.filter(status='done').count()

    return render(request, 'todo_lists.html', {'lists': lists})

@login_required
def create_todo_list(request):
    """Crea una nueva lista de tareas"""
    if request.method == 'POST':
        form = TodoListForm(request.POST)
        if form.is_valid():
            todo_list = form.save(commit=False)
            todo_list.user = request.user
            todo_list.save()
            return redirect('view_todo_list', list_id=todo_list.id)
    else:
        form = TodoListForm()
    return render(request, 'create_todo_list.html', {'form': form})

@login_required
def view_todo_list(request, list_id):
    """Muestra el detalle de una lista de tareas con sus tareas agrupadas por estado"""
    todo_list = get_object_or_404(TodoList, id=list_id)
    
    # Verificar que el usuario sea el propietario
    if todo_list.user != request.user:
        return HttpResponseForbidden("No tienes permiso para ver esta lista")
    
    # Agrupar tareas por estado
    tasks = {
        'todo': todo_list.tasks.filter(status='todo').order_by('position'),
        'progress': todo_list.tasks.filter(status='progress').order_by('position'),
        'done': todo_list.tasks.filter(status='done').order_by('position')
    }
    
    today_date = date.today()

    return render(request, 'view_todo_list.html', {
        'todo_list': todo_list,
        'tasks': tasks,
        'today_date': today_date
    })

@login_required
def add_task(request, list_id):
    """Agrega una nueva tarea a una lista"""
    todo_list = get_object_or_404(TodoList, id=list_id)
    
    # Verificar que el usuario sea el propietario
    if todo_list.user != request.user:
        return HttpResponseForbidden("No tienes permiso para modificar esta lista")
    
    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.todo_list = todo_list
            
            # Obtener el estado de la columna donde se agregará la tarea
            status = form.cleaned_data['status']
            task.status = status
            
            # Obtener la posición máxima actual para el estado
            max_position = Task.objects.filter(
                todo_list=todo_list, 
                status=status
            ).order_by('-position').first()
            
            if max_position:
                task.position = max_position.position + 1
            else:
                task.position = 0
            
            task.save()
            
            # Si es una solicitud AJAX, devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'id': task.id,
                    'title': task.title,
                    'description': task.description,
                    'status': task.status,
                    'position': task.position,
                    'due_date': task.due_date,
                })
            
            return redirect('view_todo_list', list_id=todo_list.id)
    else:
        form = TaskForm()
    
    return render(request, 'add_task.html', {
        'form': form,
        'todo_list': todo_list
    })

@login_required
def edit_task(request, task_id):
    """Edita una tarea existente"""
    task = get_object_or_404(Task, id=task_id)
    todo_list = task.todo_list
    
    # Verificar que el usuario sea el propietario
    if todo_list.user != request.user:
        return HttpResponseForbidden("No tienes permiso para modificar esta tarea")
    
    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            
            # Si es una solicitud AJAX, devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'id': task.id,
                    'title': task.title,
                    'description': task.description,
                    'status': task.status,
                    'position': task.position,
                    'due_date': task.due_date,
                })
            
            return redirect('view_todo_list', list_id=todo_list.id)
    else:
        form = TaskForm(instance=task)
    
    return render(request, 'edit_task.html', {
        'form': form,
        'task': task,
        'todo_list': todo_list
    })

@login_required
def delete_task(request, task_id):
    """Elimina una tarea"""
    task = get_object_or_404(Task, id=task_id)
    todo_list = task.todo_list
    
    # Verificar que el usuario sea el propietario
    if todo_list.user != request.user:
        return HttpResponseForbidden("No tienes permiso para eliminar esta tarea")
    
    if request.method == 'POST':
        task.delete()
        
        # Si es una solicitud AJAX, devolver JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        
        return redirect('view_todo_list', list_id=todo_list.id)
    
    return render(request, 'delete_task.html', {
        'task': task,
        'todo_list': todo_list
    })

@login_required
def delete_todo_list(request, list_id):
    """Elimina una lista de tareas"""
    todo_list = get_object_or_404(TodoList, id=list_id, user=request.user)
    
    if request.method == 'POST':
        todo_list.delete()
        return redirect('todo_lists')
    
    return render(request, 'delete_todo_list.html', {'todo_list': todo_list})

@login_required
@require_POST
def update_task_status(request, task_id):
    """API para actualizar el estado de una tarea"""
    task = get_object_or_404(Task, id=task_id)
    todo_list = task.todo_list
    
    # Verificar que el usuario sea el propietario
    if todo_list.user != request.user:
        return JsonResponse({'error': 'No tienes permiso para modificar esta tarea'}, status=403)
    
    try:
        data = json.loads(request.body)
        new_status = data.get('status')
        
        if new_status not in ['todo', 'progress', 'done']:
            return JsonResponse({'error': 'Estado no válido'}, status=400)
        
        # Actualizar el estado
        task.status = new_status
        task.save()
        
        return JsonResponse({
            'id': task.id,
            'status': task.status
        })
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_POST
def update_tasks_order(request, list_id):
    """API para actualizar el orden de las tareas"""
    todo_list = get_object_or_404(TodoList, id=list_id)
    
    # Verificar que el usuario sea el propietario
    if todo_list.user != request.user:
        return JsonResponse({'error': 'No tienes permiso para modificar esta lista'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        with transaction.atomic():
            for status, tasks in data.items():
                for i, task_id in enumerate(tasks):
                    Task.objects.filter(id=task_id, todo_list=todo_list).update(
                        status=status,
                        position=i
                    )
        
        return JsonResponse({'success': True})
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

# Vistas para Google Forms
# Modificar la vista gform_list para pasar los permisos directamente al contexto
@login_required
def gform_list(request):
    """Vista para listar todos los formularios del usuario"""
    # Obtener formularios propios
    own_forms = GForm.objects.filter(user=request.user)
    
    # Obtener formularios compartidos con el usuario
    shared_forms = GForm.objects.filter(permissions__user=request.user).distinct()
    
    # Preparar los permisos para cada formulario compartido
    shared_forms_with_permissions = []
    for form in shared_forms:
        # Añadir atributos para los permisos
        permission_type = form.get_user_permission(request.user)
        form.current_permission = permission_type
        
        # Añadir atributos para las verificaciones de permisos
        # Usar los métodos directamente para obtener los valores booleanos
        form.user_can_edit = form.can_edit(request.user)
        form.user_can_respond = form.can_respond(request.user)
        form.user_can_view = form.can_view(request.user)
        
        # Añadir flags para usar en la plantilla
        form.can_edit_flag = form.user_can_edit
        form.can_respond_flag = form.user_can_respond
        form.can_view_flag = form.user_can_view
        
        shared_forms_with_permissions.append(form)
    
    return render(request, 'forms_google/form_list.html', {
        'forms': own_forms,
        'shared_forms': shared_forms_with_permissions
    })

@login_required
def gform_create(request):
    """Vista para crear un nuevo formulario"""
    if request.method == 'POST':
        form = GFormForm(request.POST)
        if form.is_valid():
            new_form = form.save(commit=False)
            new_form.user = request.user
            new_form.save()
            messages.success(request, "Formulario creado correctamente")
            return redirect('gform_edit', form_id=new_form.id)
        else:
            # Añadir mensaje de error si el formulario no es válido
            messages.error(request, "Error al crear el formulario. Por favor, verifica los datos.")
    else:
        form = GFormForm()
    
    return render(request, 'forms_google/create_form.html', {'form': form})

@login_required
def gform_edit(request, form_id):
    """Vista para editar un formulario existente"""
    gform = get_object_or_404(GForm, id=form_id)
    
    # Verificar que el usuario tenga permisos para editar
    if not gform.can_edit(request.user):
        return HttpResponseForbidden("No tienes permiso para editar este formulario")
    
    if request.method == 'POST':
        form = GFormForm(request.POST, instance=gform)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario actualizado correctamente")
            return redirect('gform_edit', form_id=gform.id)
    else:
        form = GFormForm(instance=gform)
    
    # Obtener todas las preguntas del formulario
    questions = GQuestion.objects.filter(form=gform).order_by('position')
    
    # Obtener las preguntas del banco del usuario
    bank_questions = BankQuestion.objects.filter(created_by=request.user).order_by('-created_at')
    
    # Obtener los permisos actuales del formulario
    form_permissions = FormPermission.objects.filter(form=gform).select_related('user')
    
    # Obtener los enlaces de compartir
    share_links = FormShareLink.objects.filter(form=gform)
    
    # Formularios para compartir
    permission_form = FormPermissionForm()
    share_link_form = FormShareLinkForm()
    
    return render(request, 'forms_google/edit_form.html', {
        'form': form,
        'gform': gform,
        'questions': questions,
        'bank_questions': bank_questions,
        'question_form': GQuestionForm(),
        'option_form': GOptionForm(),
        'form_permissions': form_permissions,
        'share_links': share_links,
        'permission_form': permission_form,
        'share_link_form': share_link_form,
        'is_owner': request.user == gform.user,
    })

@login_required
def gform_share_form(request, form_id):
    """Vista para gestionar los permisos de un formulario"""
    gform = get_object_or_404(GForm, id=form_id)
    
    # Verificar que el usuario sea el propietario
    if gform.user != request.user:
        return HttpResponseForbidden("Solo el propietario puede gestionar permisos")
    
    # Obtener los permisos actuales del formulario
    form_permissions = FormPermission.objects.filter(form=gform).select_related('user')
    
    # Obtener los enlaces de compartir
    share_links = FormShareLink.objects.filter(form=gform)
    
    # Formularios para compartir
    permission_form = FormPermissionForm()
    share_link_form = FormShareLinkForm()
    
    return render(request, 'forms_google/share_form.html', {
        'form': gform,
        'form_permissions': form_permissions,
        'share_links': share_links,
        'permission_form': permission_form,
        'share_link_form': share_link_form,
    })

@login_required
def gform_delete(request, form_id):
    """Vista para eliminar un formulario"""
    gform = get_object_or_404(GForm, id=form_id)
    
    # Verificar que el usuario sea el propietario
    if gform.user != request.user:
        return HttpResponseForbidden("No tienes permiso para eliminar este formulario")
    
    if request.method == 'POST':
        # Eliminar el formulario sin afectar al banco de preguntas
        gform.delete()
        messages.success(request, "Formulario eliminado correctamente")
        return redirect('gform_list')
    
    return render(request, 'forms_google/delete_form.html', {'gform': gform})

@login_required
@require_POST
def gform_add_question(request, form_id):
    """Vista para añadir una pregunta al formulario"""
    gform = get_object_or_404(GForm, id=form_id)
    
    # Verificar que el usuario tenga permisos para editar
    if not gform.can_edit(request.user):
        return HttpResponseForbidden("No tienes permiso para modificar este formulario")
    
    form = GQuestionForm(request.POST, request.FILES)
    if form.is_valid():
        question = form.save(commit=False)
        question.form = gform
        
        # Asegúrate de que el campo allow_attachments se guarde correctamente
        question.allow_attachments = 'allow_attachments' in request.POST
        
        # Obtener la posición máxima actual
        max_position = GQuestion.objects.filter(form=gform).order_by('-position').first()
        if max_position:
            question.position = max_position.position + 1
        else:
            question.position = 0
        
        # Guardar en el banco de preguntas si se seleccionó
        if 'in_question_bank' in request.POST:
            # Crear una nueva pregunta en el banco
            bank_question = BankQuestion(
                text=question.text,
                question_type=question.question_type,
                help_text=question.help_text,
                is_required=question.is_required,
                min_value=question.min_value,
                max_value=question.max_value,
                min_label=question.min_label,
                max_label=question.max_label,
                allow_attachments=question.allow_attachments,
                created_by=request.user
            )
            
            # Copiar la imagen si existe
            if 'image' in request.FILES:
                bank_question.image = request.FILES['image']
                
            bank_question.save()
            
            # Vincular la pregunta del formulario con la del banco
            question.bank_question = bank_question
        
        question.save()
        
        # Si la pregunta es de tipo selección y se guardó en el banco, copiar las opciones al banco
        if question.bank_question and question.question_type in ['multiple_choice', 'checkbox', 'dropdown']:
            # Primero necesitamos guardar las opciones para la pregunta del formulario
            options_data = request.POST.getlist('option_text', [])
            for i, option_text in enumerate(options_data):
                if option_text.strip():
                    option = GOption.objects.create(
                        question=question,
                        text=option_text.strip(),
                        position=i
                    )
                    
                    # Crear la opción correspondiente en el banco
                    BankOption.objects.create(
                        question=question.bank_question,
                        text=option.text,
                        position=option.position
                    )
        
        # Si es una solicitud AJAX, devolver JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'id': question.id,
                'text': question.text,
                'type': question.get_question_type_display()
            })
        
        messages.success(request, "Pregunta añadida correctamente")
        return redirect('gform_edit', form_id=gform.id)
    
    # Si hay errores en el formulario
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    
    messages.error(request, "Error al añadir la pregunta")
    return redirect('gform_edit', form_id=gform.id)

@login_required
@require_POST
def gform_edit_question(request, question_id):
    """Vista para editar una pregunta existente"""
    question = get_object_or_404(GQuestion, id=question_id)
    gform = question.form
    
    # Verificar que el usuario tenga permisos para editar
    if not gform.can_edit(request.user):
        return HttpResponseForbidden("No tienes permiso para modificar este formulario")
    
    form = GQuestionForm(request.POST, request.FILES, instance=question)
    if form.is_valid():
        question = form.save(commit=False)
        # Asegúrate de que el campo allow_attachments se guarde correctamente
        question.allow_attachments = 'allow_attachments' in request.POST
        
        # Actualizar el banco de preguntas si existe
        if question.bank_question:
            bank_question = question.bank_question
            bank_question.text = question.text
            bank_question.question_type = question.question_type
            bank_question.help_text = question.help_text
            bank_question.is_required = question.is_required
            bank_question.min_value = question.min_value
            bank_question.max_value = question.max_value
            bank_question.min_label = question.min_label
            bank_question.max_label = question.max_label
            bank_question.allow_attachments = question.allow_attachments
            
            # Actualizar la imagen si se proporciona
            if 'image' in request.FILES:
                bank_question.image = request.FILES['image']
                
            bank_question.save()
        # Si no está en el banco y se seleccionó la opción, crear una nueva pregunta en el banco
        elif 'in_question_bank' in request.POST:
            bank_question = BankQuestion(
                text=question.text,
                question_type=question.question_type,
                help_text=question.help_text,
                is_required=question.is_required,
                min_value=question.min_value,
                max_value=question.max_value,
                min_label=question.min_label,
                max_label=question.max_label,
                allow_attachments=question.allow_attachments,
                created_by=request.user
            )
            
            # Copiar la imagen si existe
            if 'image' in request.FILES:
                bank_question.image = request.FILES['image']
                
            bank_question.save()
            
            # Vincular la pregunta del formulario con la del banco
            question.bank_question = bank_question
            
            # Si la pregunta es de tipo selección, copiar las opciones al banco
            if question.question_type in ['multiple_choice', 'checkbox', 'dropdown']:
                for option in question.options.all():
                    BankOption.objects.create(
                        question=bank_question,
                        text=option.text,
                        position=option.position
                    )
        
        question.save()
        
        # Si es una solicitud AJAX, devolver JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'id': question.id,
                'text': question.text,
                'type': question.get_question_type_display()
            })
        
        messages.success(request, "Pregunta actualizada correctamente")
        return redirect('gform_edit', form_id=gform.id)
    
    # Si hay errores en el formulario
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    
    messages.error(request, "Error al actualizar la pregunta")
    return redirect('gform_edit', form_id=gform.id)

@login_required
@require_POST
def gform_delete_question(request, question_id):
    """Vista para eliminar una pregunta"""
    try:
        # Intentar obtener la pregunta por ID
        question = get_object_or_404(GQuestion, id=question_id)
        gform = question.form
        
        # Verificar que el usuario tenga permisos para editar
        if not gform.can_edit(request.user):
            return HttpResponseForbidden("No tienes permiso para modificar este formulario")
        
        # Guardar el ID del formulario antes de eliminar la pregunta
        form_id = gform.id
        
        # Eliminar la pregunta del formulario sin importar si está vinculada al banco
        question.delete()
        
        # Si es una solicitud AJAX, devolver JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'Pregunta eliminada correctamente'})
        
        messages.success(request, "Pregunta eliminada correctamente del formulario")
        return redirect('gform_edit', form_id=form_id)
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"Error al eliminar pregunta: {str(e)}\n{error_traceback}")
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': str(e)})
        
        # Si no podemos obtener el ID del formulario de la pregunta (porque no existe),
        # intentamos obtenerlo de la URL de referencia
        try:
            referer = request.META.get('HTTP_REFERER', '')
            import re
            form_id_match = re.search(r'/forms/(\d+)/edit/', referer)
            if form_id_match:
                form_id = form_id_match.group(1)
                messages.error(request, f"Error al eliminar la pregunta: {str(e)}")
                return redirect('gform_edit', form_id=form_id)
        except:
            pass
        
        # Si todo falla, redirigir a la lista de formularios
        messages.error(request, f"Error al eliminar la pregunta: {str(e)}")
        return redirect('gform_list')

@login_required
@require_POST
def gform_add_option(request, question_id):
    """Vista para añadir una opción a una pregunta"""
    question = get_object_or_404(GQuestion, id=question_id)
    gform = question.form
    
    # Verificar que el usuario tenga permisos para editar
    if not gform.can_edit(request.user):
        return HttpResponseForbidden("No tienes permiso para modificar este formulario")
    
    # Verificar que la pregunta sea de un tipo que admite opciones
    if question.question_type not in ['multiple_choice', 'checkbox', 'dropdown']:
        return HttpResponseForbidden("Este tipo de pregunta no admite opciones")
    
    form = GOptionForm(request.POST)
    if form.is_valid():
        option = form.save(commit=False)
        option.question = question
        
        # Obtener la posición máxima actual
        max_position = GOption.objects.filter(question=question).order_by('-position').first()
        if max_position:
            option.position = max_position.position + 1
        else:
            option.position = 0
        
        option.save()
        
        # Si la pregunta está vinculada al banco, añadir también la opción al banco
        if question.bank_question:
            BankOption.objects.create(
                question=question.bank_question,
                text=option.text,
                position=option.position
            )
        
        # Si es una solicitud AJAX, devolver JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'id': option.id,
                'text': option.text
            })
        
        messages.success(request, "Opción añadida correctamente")
        return redirect('gform_edit', form_id=gform.id)
    
    # Si hay errores en el formulario
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    
    messages.error(request, "Error al añadir la opción")
    return redirect('gform_edit', form_id=gform.id)

@login_required
@require_POST
def gform_edit_option(request, option_id):
    """Vista para editar una opción existente"""
    option = get_object_or_404(GOption, id=option_id)
    question = option.question
    gform = question.form
    
    # Verificar que el usuario tenga permisos para editar
    if not gform.can_edit(request.user):
        return HttpResponseForbidden("No tienes permiso para modificar este formulario")
    
    form = GOptionForm(request.POST, instance=option)
    if form.is_valid():
        option = form.save()
        
        # Si la pregunta está vinculada al banco, actualizar también la opción en el banco
        if question.bank_question:
            # Buscar la opción correspondiente en el banco por posición
            try:
                bank_option = BankOption.objects.get(
                    question=question.bank_question,
                    position=option.position
                )
                bank_option.text = option.text
                bank_option.save()
            except BankOption.DoesNotExist:
                # Si no existe, crearla
                BankOption.objects.create(
                    question=question.bank_question,
                    text=option.text,
                    position=option.position
                )
        
        # Si es una solicitud AJAX, devolver JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'id': option.id,
                'text': option.text
            })
        
        messages.success(request, "Opción actualizada correctamente")
        return redirect('gform_edit', form_id=gform.id)
    
    # Si hay errores en el formulario
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    
    messages.error(request, "Error al actualizar la opción")
    return redirect('gform_edit', form_id=gform.id)

@login_required
@require_POST
def gform_delete_option(request, option_id):
    """Vista para eliminar una opción"""
    option = get_object_or_404(GOption, id=option_id)
    question = option.question
    gform = question.form
    
    # Verificar que el usuario tenga permisos para editar
    if not gform.can_edit(request.user):
        return HttpResponseForbidden("No tienes permiso para modificar este formulario")
    
    # Si la pregunta está vinculada al banco, eliminar también la opción en el banco
    if question.bank_question:
        # Buscar la opción correspondiente en el banco por posición
        try:
            bank_option = BankOption.objects.get(
                question=question.bank_question,
                position=option.position
            )
            bank_option.delete()
        except BankOption.DoesNotExist:
            pass
    
    option.delete()
    
    # Si es una solicitud AJAX, devolver JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    messages.success(request, "Opción eliminada correctamente")
    return redirect('gform_edit', form_id=gform.id)

@login_required
@require_POST
def gform_update_question_order(request, form_id):
    """Vista para actualizar el orden de las preguntas"""
    gform = get_object_or_404(GForm, id=form_id)
    
    # Verificar que el usuario tenga permisos para editar
    if not gform.can_edit(request.user):
        return JsonResponse({'error': 'No tienes permiso para modificar este formulario'}, status=403)
    
    try:
        data = json.loads(request.body)
        question_ids = data.get('question_ids', [])
        
        with transaction.atomic():
            for i, question_id in enumerate(question_ids):
                GQuestion.objects.filter(id=question_id, form=gform).update(position=i)
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required
@require_POST
def gform_update_option_order(request, question_id):
    """Vista para actualizar el orden de las opciones"""
    question = get_object_or_404(GQuestion, id=question_id)
    gform = question.form
    
    # Verificar que el usuario tenga permisos para editar
    if not gform.can_edit(request.user):
        return JsonResponse({'error': 'No tienes permiso para modificar este formulario'}, status=403)
    
    try:
        data = json.loads(request.body)
        option_ids = data.get('option_ids', [])
        
        with transaction.atomic():
            for i, option_id in enumerate(option_ids):
                GOption.objects.filter(id=option_id, question=question).update(position=i)
                
                # Si la pregunta está vinculada al banco, actualizar también el orden en el banco
                if question.bank_question:
                    option = GOption.objects.get(id=option_id)
                    # Buscar la opción correspondiente en el banco por texto
                    try:
                        bank_option = BankOption.objects.get(
                            question=question.bank_question,
                            text=option.text
                        )
                        bank_option.position = i
                        bank_option.save()
                    except BankOption.DoesNotExist:
                        pass
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required
def gform_toggle_publish(request, form_id):
    """Vista para publicar/despublicar un formulario"""
    gform = get_object_or_404(GForm, id=form_id)
    
    # Verificar que el usuario sea el propietario o tenga permisos de editor
    if not gform.can_edit(request.user):
        return HttpResponseForbidden("No tienes permiso para modificar este formulario")
    
    gform.is_published = not gform.is_published
    gform.save()
    
    status = "publicado" if gform.is_published else "despublicado"
    messages.success(request, f"Formulario {status} correctamente")
    
    return redirect('gform_edit', form_id=gform.id)

def gform_view(request, form_id):
    """Vista para ver y responder a un formulario"""
    gform = get_object_or_404(GForm, id=form_id)
    
    # Verificar que el formulario esté publicado o que el usuario tenga permisos para verlo
    if not gform.is_published and (request.user.is_anonymous or not gform.can_view(request.user)):
        return HttpResponseForbidden("Este formulario no está disponible")
    
    questions = GQuestion.objects.filter(form=gform).order_by('position')
    
    # Crear contexto para las preguntas de tipo escala lineal
    for question in questions:
        if question.question_type == 'linear_scale':
            min_val = question.min_value or 1
            max_val = question.max_value or 5
            question.range_values = range(min_val, max_val + 1)
    
    if request.method == 'POST':
        # Procesar la respuesta al formulario
        if not gform.is_published and not gform.can_respond(request.user):
            return HttpResponseForbidden("Este formulario no está aceptando respuestas")
        
        # Crear una nueva respuesta
        response = GResponse.objects.create(
            form=gform,
            respondent=request.user if request.user.is_authenticated else None,
            respondent_email=request.POST.get('email', None)
        )
        
        # Procesar cada pregunta
        for question in questions:
            field_name = f'question_{question.id}'
            file_field_name = f'file_{question.id}'
            url_field_name = f'url_{question.id}'
            
            if field_name in request.POST or file_field_name in request.FILES or url_field_name in request.POST:
                answer = GAnswer.objects.create(
                    response=response,
                    question=question
                )
                
                if field_name in request.POST:
                    if question.question_type in ['short_text', 'paragraph', 'date', 'time', 'linear_scale']:
                        answer.text_answer = request.POST.get(field_name)
                        answer.save()
                    
                    elif question.question_type in ['multiple_choice', 'dropdown']:
                        option_id = request.POST.get(field_name)
                        if option_id:
                            option = get_object_or_404(GOption, id=option_id)
                            GSelectedOption.objects.create(answer=answer, option=option)
                    
                    elif question.question_type == 'checkbox':
                        # Para checkbox, value puede ser una lista
                        option_ids = request.POST.getlist(field_name)
                        for option_id in option_ids:
                            option = get_object_or_404(GOption, id=option_id)
                            GSelectedOption.objects.create(answer=answer, option=option)
                
                # Procesar archivos adjuntos solo si la pregunta permite adjuntos
                if question.allow_attachments:
                    if file_field_name in request.FILES:
                        file = request.FILES[file_field_name]
                        if file.content_type.startswith('image/'):
                            answer.image = file
                        elif file.content_type.startswith('video/'):
                            answer.video = file
                        answer.save()
                    
                    # Procesar URL
                    if url_field_name in request.POST and request.POST.get(url_field_name):
                        answer.file_url = request.POST.get(url_field_name)
                        answer.save()
        
        messages.success(request, "¡Gracias por tu respuesta!")
        return redirect('gform_thank_you', form_id=gform.id)
    
    return render(request, 'forms_google/view_form.html', {
        'gform': gform,
        'questions': questions,
        'is_owner': request.user == gform.user,
        'can_edit': gform.can_edit(request.user) if request.user.is_authenticated else False,
    })

def gform_respond(request, form_id):
    """Vista exclusiva para responder a un formulario sin opciones de edición"""
    gform = get_object_or_404(GForm, id=form_id)
    
    # Verificar que el formulario esté publicado o el usuario tenga permisos para responder
    if not gform.is_published and (request.user.is_anonymous or not gform.can_respond(request.user)):
        messages.warning(request, "Este formulario no está disponible para respuestas")
        return redirect('gform_list')
    
    questions = GQuestion.objects.filter(form=gform).order_by('position')
    
    # Verificar que el formulario tenga preguntas
    if not questions.exists():
        messages.warning(request, "Este formulario no tiene preguntas. No se puede responder.")
        return redirect('gform_list')
    
    # Crear contexto para las preguntas de tipo escala lineal
    for question in questions:
        if question.question_type == 'linear_scale':
            min_val = question.min_value or 1
            max_val = question.max_value or 5
            question.range_values = range(min_val, max_val + 1)
    
    if request.method == 'POST':
        # Procesar la respuesta al formulario
        
        # Crear una nueva respuesta
        response = GResponse.objects.create(
            form=gform,
            respondent=request.user if request.user.is_authenticated else None,
            respondent_email=request.POST.get('email', None)
        )
        
        # Procesar cada pregunta
        for question in questions:
            field_name = f'question_{question.id}'
            file_field_name = f'file_{question.id}'
            url_field_name = f'url_{question.id}'
            
            if field_name in request.POST or file_field_name in request.FILES or url_field_name in request.POST:
                answer = GAnswer.objects.create(
                    response=response,
                    question=question
                )
                
                if field_name in request.POST:
                    if question.question_type in ['short_text', 'paragraph', 'date', 'time', 'linear_scale']:
                        answer.text_answer = request.POST.get(field_name)
                        answer.save()
                    
                    elif question.question_type in ['multiple_choice', 'dropdown']:
                        option_id = request.POST.get(field_name)
                        if option_id:
                            option = get_object_or_404(GOption, id=option_id)
                            GSelectedOption.objects.create(answer=answer, option=option)
                    
                    elif question.question_type == 'checkbox':
                        # Para checkbox, value puede ser una lista
                        option_ids = request.POST.getlist(field_name)
                        for option_id in option_ids:
                            option = get_object_or_404(GOption, id=option_id)
                            GSelectedOption.objects.create(answer=answer, option=option)
                
                # Procesar archivos adjuntos solo si la pregunta permite adjuntos
                if question.allow_attachments:
                    if file_field_name in request.FILES:
                        file = request.FILES[file_field_name]
                        if file.content_type.startswith('image/'):
                            answer.image = file
                        elif file.content_type.startswith('video/'):
                            answer.video = file
                        answer.save()
                    
                    # Procesar URL
                    if url_field_name in request.POST and request.POST.get(url_field_name):
                        answer.file_url = request.POST.get(url_field_name)
                        answer.save()
        
        messages.success(request, "¡Gracias por tu respuesta!")
        return redirect('gform_thank_you', form_id=gform.id)
    
    return render(request, 'forms_google/respond_form.html', {
        'gform': gform,
        'questions': questions
    })

def gform_thank_you(request, form_id):
    """Vista de agradecimiento después de enviar una respuesta"""
    gform = get_object_or_404(GForm, id=form_id)
    return render(request, 'forms_google/thank_you.html', {'gform': gform})

@login_required
def gform_responses(request, form_id):
    """Vista para ver las respuestas a un formulario"""
    gform = get_object_or_404(GForm, id=form_id)
    
    # Verificar que el usuario tenga permisos para ver las respuestas
    if not gform.can_edit(request.user):
        return HttpResponseForbidden("No tienes permiso para ver las respuestas de este formulario")
    
    responses = GResponse.objects.filter(form=gform).order_by('-created_at')
    
    # Si es una solicitud AJAX para exportar a CSV, devolver JSON con todos los datos
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.GET.get('format') == 'json':
        # Obtener todas las preguntas
        questions = GQuestion.objects.filter(form=gform).order_by('position')
        questions_data = []
        
        for question in questions:
            question_data = {
                'id': question.id,
                'text': question.text,
                'question_type': question.question_type,
                'is_required': question.is_required
            }
            questions_data.append(question_data)
        
        # Obtener todas las respuestas con sus detalles
        responses_data = []
        
        for response in responses:
            response_data = {
                'id': response.id,
                'respondent': response.respondent.username if response.respondent else None,
                'respondent_email': response.respondent_email,
                'created_at': response.created_at.strftime('%d/%m/%Y %H:%M'),
                'answers': []
            }
            
            # Obtener respuestas para cada pregunta
            answers = GAnswer.objects.filter(response=response).select_related('question')
            
            for answer in answers:
                answer_data = {
                    'question_id': answer.question.id,
                    'text_answer': answer.text_answer,
                    'image': answer.image.url if answer.image else None,
                    'video': answer.video.url if answer.video else None,
                    'file_url': answer.file_url,
                    'selected_options': []
                }
                
                # Obtener opciones seleccionadas
                selected_options = GSelectedOption.objects.filter(answer=answer).select_related('option')
                
                for selected in selected_options:
                    answer_data['selected_options'].append({
                        'id': selected.option.id,
                        'text': selected.option.text
                    })
                
                response_data['answers'].append(answer_data)
            
            responses_data.append(response_data)
        
        return JsonResponse({
            'form': {
                'id': gform.id,
                'title': gform.title,
                'description': gform.description
            },
            'questions': questions_data,
            'responses': responses_data
        })
    
    return render(request, 'forms_google/view_responses.html', {
        'gform': gform,
        'responses': responses,
        'is_owner': request.user == gform.user,
    })

@login_required
def gform_response_detail(request, response_id):
    """Vista para ver el detalle de una respuesta"""
    response = get_object_or_404(GResponse, id=response_id)
    gform = response.form
    
    # Verificar que el usuario tenga permisos para ver las respuestas
    if not gform.can_edit(request.user):
        return HttpResponseForbidden("No tienes permiso para ver esta respuesta")
    
    # Obtener todas las preguntas y respuestas
    questions = GQuestion.objects.filter(form=gform).order_by('position')
    answers = GAnswer.objects.filter(response=response).select_related('question')
    
    # Crear un diccionario para facilitar el acceso a las respuestas
    answer_dict = {}
    for answer in answers:
        answer_dict[answer.question.id] = answer
    
    # Si es una solicitud AJAX para exportar a CSV, devolver JSON con los datos
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.GET.get('format') == 'json':
        # Preparar datos para exportación
        response_data = {
            'id': response.id,
            'respondent': response.respondent.username if response.respondent else None,
            'respondent_email': response.respondent_email,
            'created_at': response.created_at.strftime('%d/%m/%Y %H:%M'),
            'answers': []
        }
        
        # Obtener respuestas para cada pregunta
        for question in questions:
            answer_data = {
                'question_id': question.id,
                'question_text': question.text,
                'question_type': question.question_type,
                'text_answer': None,
                'image': None,
                'video': None,
                'file_url': None,
                'selected_options': []
            }
            
            if question.id in answer_dict:
                answer = answer_dict[question.id]
                answer_data['text_answer'] = answer.text_answer
                answer_data['image'] = answer.image.url if answer.image else None
                answer_data['video'] = answer.video.url if answer.video else None
                answer_data['file_url'] = answer.file_url
                
                # Obtener opciones seleccionadas
                selected_options = GSelectedOption.objects.filter(answer=answer).select_related('option')
                for selected in selected_options:
                    answer_data['selected_options'].append({
                        'id': selected.option.id,
                        'text': selected.option.text
                    })
            
            response_data['answers'].append(answer_data)
        
        return JsonResponse({
            'form': {
                'id': gform.id,
                'title': gform.title,
                'description': gform.description
            },
            'response': response_data
        })
    
    return render(request, 'forms_google/response_detail.html', {
        'gform': gform,
        'response': response,
        'questions': questions,
        'answer_dict': answer_dict,
        'is_owner': request.user == gform.user,
    })

@login_required
def gform_response_data(request, response_id):
    """Vista para obtener los datos de una respuesta en formato JSON"""
    response = get_object_or_404(GResponse, id=response_id)
    gform = response.form
    
    # Verificar que el usuario tenga permisos para ver las respuestas
    if not gform.can_edit(request.user):
        return JsonResponse({"error": "No tienes permiso para ver esta respuesta"}, status=403)
    
    # Obtener todas las preguntas y respuestas
    questions = GQuestion.objects.filter(form=gform).order_by('position')
    answers = GAnswer.objects.filter(response=response).select_related('question')
    
    # Crear un diccionario para facilitar el acceso a las respuestas
    answer_dict = {}
    for answer in answers:
        answer_dict[answer.question.id] = answer
    
    # Preparar datos para exportación
    response_data = {
        'id': response.id,
        'respondent': response.respondent.username if response.respondent else None,
        'respondent_email': response.respondent_email,
        'created_at': response.created_at.strftime('%d/%m/%Y %H:%M'),
        'answers': []
    }
    
    # Obtener respuestas para cada pregunta
    for question in questions:
        answer_data = {
            'question_id': question.id,
            'question_text': question.text,
            'question_type': question.question_type,
            'text_answer': None,
            'image': None,
            'video': None,
            'file_url': None,
            'selected_options': []
        }
        
        if question.id in answer_dict:
            answer = answer_dict[question.id]
            answer_data['text_answer'] = answer.text_answer
            answer_data['image'] = answer.image.url if answer.image else None
            answer_data['video'] = answer.video.url if answer.video else None
            answer_data['file_url'] = answer.file_url
            
            # Obtener opciones seleccionadas
            selected_options = GSelectedOption.objects.filter(answer=answer).select_related('option')
            for selected in selected_options:
                answer_data['selected_options'].append({
                    'id': selected.option.id,
                    'text': selected.option.text
                })
        
        response_data['answers'].append(answer_data)
    
    return JsonResponse({
        'form': {
            'id': gform.id,
            'title': gform.title,
            'description': gform.description
        },
        'response': response_data
    })

@login_required
def export_response_to_excel(request, response_id):
    """Vista para exportar una respuesta a Excel con imágenes incrustadas"""
    response_obj = get_object_or_404(GResponse, id=response_id)
    gform = response_obj.form
    
    # Verificar que el usuario tenga permisos para ver las respuestas
    if not gform.can_edit(request.user):
        return HttpResponse("No tienes permiso para ver esta respuesta", status=403)
    
    # Obtener todas las preguntas y respuestas
    questions = GQuestion.objects.filter(form=gform).order_by('position')
    answers = GAnswer.objects.filter(response=response_obj).select_related('question')
    
    # Crear un diccionario para facilitar el acceso a las respuestas
    answer_dict = {}
    for answer in answers:
        answer_dict[answer.question.id] = answer
    
    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Respuesta #{response_obj.id}"
    
    # Añadir información del formulario
    ws['A1'] = "Formulario:"
    ws['B1'] = gform.title
    ws['A2'] = "Descripción:"
    ws['B2'] = gform.description or ""
    ws['A3'] = "Respondente:"
    ws['B3'] = response_obj.respondent.username if response_obj.respondent else (response_obj.respondent_email or "Anónimo")
    ws['A4'] = "Fecha de respuesta:"
    ws['B4'] = response_obj.created_at.strftime('%d/%m/%Y %H:%M')
    
    # Crear respuesta HTTP con el archivo Excel
    http_response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    http_response['Content-Disposition'] = f'attachment; filename=respuesta_{response_obj.id}_{gform.title.replace(" ", "_")}.xlsx'
    
    # Guardar el libro de Excel en la respuesta HTTP
    wb.save(http_response)
    
    return http_response

# Nuevas vistas para el banco de preguntas independiente

# Modificar la función gform_save_question_to_bank para usar el nuevo sistema de hash
@login_required
@require_POST
def gform_save_question_to_bank(request, question_id):
    """Vista para guardar una pregunta en el banco de preguntas"""
    question = get_object_or_404(GQuestion, id=question_id)
    gform = question.form
    
    # Verificar que el usuario tenga permisos para editar
    if not gform.can_edit(request.user):
        return HttpResponseForbidden("No tienes permiso para modificar este formulario")
    
    try:
        # Generar el hash para esta pregunta
        content = f"{question.text}|{question.question_type}|{question.help_text}|{question.is_required}|{request.user.id}"
        question_hash = hashlib.sha256(content.encode()).hexdigest()
        
        # Verificar si ya existe una pregunta con el mismo hash
        existing_question = BankQuestion.objects.filter(
            created_by=request.user,
            question_hash=question_hash
        ).first()
        
        if existing_question:
            # Si ya existe, usar esa en lugar de crear una nueva
            bank_question = existing_question
            message = "La pregunta ya existía en el banco y ha sido recuperada"
        else:
            # Crear una nueva pregunta en el banco
            bank_question = BankQuestion(
                text=question.text,
                question_type=question.question_type,
                help_text=question.help_text,
                is_required=question.is_required,
                min_value=question.min_value,
                max_value=question.max_value,
                min_label=question.min_label,
                max_label=question.max_label,
                allow_attachments=question.allow_attachments,
                created_by=request.user
            )
            
            # Copiar la imagen si existe
            if question.image:
                bank_question.image = question.image
                
            bank_question.save()
            message = "Pregunta guardada en el banco correctamente"
        
        # Si la pregunta tiene opciones, verificar si ya existen o crearlas
        if question.question_type in ['multiple_choice', 'checkbox', 'dropdown']:
            # Primero eliminar opciones existentes para evitar duplicados
            if not existing_question:
                for option in question.options.all():
                    # Verificar si ya existe una opción con el mismo texto
                    existing_option = BankOption.objects.filter(
                        question=bank_question,
                        text=option.text
                    ).first()
                    
                    if not existing_option:
                        BankOption.objects.create(
                            question=bank_question,
                            text=option.text,
                            position=option.position
                        )
        
        return JsonResponse({
            'success': True,
            'message': message,
            'bank_question_id': str(bank_question.id)
        })
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"Error al guardar pregunta en banco: {str(e)}\n{error_traceback}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def gform_question_bank(request):
    """Vista para ver el banco de preguntas del usuario"""
    bank_questions = BankQuestion.objects.filter(created_by=request.user).order_by('-created_at')
    
    return render(request, 'forms_google/question_bank.html', {
        'bank_questions': bank_questions
    })

# Modificar la función gform_question_bank_json para usar el nuevo sistema de hash
@login_required
def gform_question_bank_json(request):
    """Vista para obtener las preguntas del banco del usuario en formato JSON"""
    try:
        # Obtener preguntas del banco sin duplicados usando el hash
        bank_questions = BankQuestion.objects.filter(created_by=request.user).order_by('-created_at')
        
        # Manejar el caso cuando no hay preguntas
        if not bank_questions.exists():
            return JsonResponse({
                'success': True,
                'bank_questions': []
            })
        
        # Preparar los datos para la respuesta JSON
        questions_data = []
        
        for question in bank_questions:
            questions_data.append({
                'id': str(question.id),
                'text': question.text,
                'question_type': question.question_type,
                'question_type_display': question.get_question_type_display(),
                'help_text': question.help_text,
                'usage_count': question.usage_count,
                'is_required': question.is_required,
                'allow_attachments': question.allow_attachments,
                'created_at': question.created_at.isoformat() if question.created_at else None,
                'hash': question.question_hash[:8] if question.question_hash else None  # Solo para depuración
            })
        
        return JsonResponse({
            'success': True,
            'bank_questions': questions_data
        })
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"Error en gform_question_bank_json: {str(e)}\n{error_traceback}")
        return JsonResponse({
            'success': False,
            'error': str(e),
            'traceback': error_traceback
        }, status=500)

@login_required
def gform_question_bank_detail(request, question_id):
    """Vista para obtener los detalles de una pregunta del banco"""
    question = get_object_or_404(BankQuestion, id=question_id, created_by=request.user)
    
    # Si es una solicitud AJAX, devolver JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'question': {
                'id': str(question.id),
                'text': question.text,
                'question_type': question.question_type,
                'question_type_display': question.get_question_type_display(),
                'help_text': question.help_text,
                'is_required': question.is_required,
                'allow_attachments': question.allow_attachments,
                'min_value': question.min_value,
                'max_value': question.max_value,
                'min_label': question.min_label,
                'max_label': question.max_label,
                'usage_count': question.usage_count,
            }
        })
    
    # Para solicitudes normales, redirigir al banco de preguntas
    return redirect('gform_question_bank')

@login_required
@require_POST
def gform_question_bank_edit(request, question_id):
    """Vista para editar una pregunta del banco"""
    question = get_object_or_404(BankQuestion, id=question_id, created_by=request.user)
    
    try:
        # Actualizar campos básicos
        question.text = request.POST.get('text', question.text)
        question.help_text = request.POST.get('help_text', question.help_text)
        question.question_type = request.POST.get('question_type', question.question_type)
        question.is_required = request.POST.get('is_required', '') == 'on'
        question.allow_attachments = request.POST.get('allow_attachments', '') == 'on'
        
        # Actualizar campos específicos para escala lineal
        if question.question_type == 'linear_scale':
            question.min_value = request.POST.get('min_value', question.min_value)
            question.max_value = request.POST.get('max_value', question.max_value)
            question.min_label = request.POST.get('min_label', question.min_label)
            question.max_label = request.POST.get('max_label', question.max_label)
        
        # Actualizar imagen si se proporciona
        if 'image' in request.FILES:
            question.image = request.FILES['image']
        
        question.save()
        
        # Actualizar también todas las preguntas de formularios vinculadas a esta pregunta del banco
        for form_question in question.form_questions.all():
            form_question.text = question.text
            form_question.help_text = question.help_text
            form_question.question_type = question.question_type
            form_question.is_required = question.is_required
            form_question.allow_attachments = question.allow_attachments
            
            if question.question_type == 'linear_scale':
                form_question.min_value = question.min_value
                form_question.max_value = question.max_value
                form_question.min_label = question.min_label
                form_question.max_label = question.max_label
            
            if 'image' in request.FILES:
                form_question.image = question.image
            
            form_question.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Pregunta actualizada correctamente'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@require_POST
def gform_question_bank_delete(request, question_id):
    """Vista para eliminar una pregunta del banco"""
    question = get_object_or_404(BankQuestion, id=question_id, created_by=request.user)
    
    try:
        # Desvincula todas las preguntas de formularios que usan esta pregunta del banco
        for form_question in question.form_questions.all():
            form_question.bank_question = None
            form_question.save()
        
        # Elimina la pregunta del banco
        question.delete()
        
        # Verificar si hay otras preguntas con el mismo ID (por si acaso)
        duplicate_questions = BankQuestion.objects.filter(id=question_id)
        if duplicate_questions.exists():
            for dupe in duplicate_questions:
                dupe.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Pregunta eliminada correctamente del banco'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

# Modificar la función gform_add_from_bank para usar el nuevo sistema de hash
@login_required
@require_POST
def gform_add_from_bank(request, form_id, question_id):
    """Vista para añadir una pregunta del banco a un formulario"""
    try:
        form = get_object_or_404(GForm, id=form_id)
        bank_question = get_object_or_404(BankQuestion, id=question_id)
        
        # Verificar que el usuario tenga permisos para editar
        if not form.can_edit(request.user):
            return HttpResponseForbidden("No tienes permiso para modificar este formulario")
        
        # Determinar el orden de la nueva pregunta
        max_position = GQuestion.objects.filter(form=form).aggregate(models.Max('position'))['position__max'] or 0
        
        # Crear la nueva pregunta en el formulario como una copia independiente
        form_question = GQuestion(
            form=form,
            text=bank_question.text,
            help_text=bank_question.help_text,
            question_type=bank_question.question_type,
            is_required=bank_question.is_required,
            allow_attachments=bank_question.allow_attachments,
            min_value=bank_question.min_value,
            max_value=bank_question.max_value,
            min_label=bank_question.min_label,
            max_label=bank_question.max_label,
            position=max_position + 1,
            bank_question=bank_question  # Mantener la referencia para seguimiento
        )
        
        # Copiar la imagen si existe
        if bank_question.image:
            form_question.image = bank_question.image
            
        form_question.save()
        
        # Si la pregunta tiene opciones, copiarlas al formulario
        options_data = []
        if bank_question.question_type in ['multiple_choice', 'checkbox', 'dropdown']:
            for bank_option in bank_question.options.all():
                option = GOption.objects.create(
                    question=form_question,
                    text=bank_option.text,
                    position=bank_option.position
                )
                options_data.append({
                    'id': option.id,
                    'text': option.text,
                    'position': option.position
                })
        
        # Preparar datos de la pregunta para la respuesta JSON
        question_data = {
            'id': form_question.id,
            'text': form_question.text,
            'help_text': form_question.help_text,
            'question_type': form_question.question_type,
            'question_type_display': form_question.get_question_type_display(),
            'is_required': form_question.is_required,
            'allow_attachments': form_question.allow_attachments,
            'in_question_bank': True,  # Ahora sí está vinculada
            'position': form_question.position,
            'image': form_question.image.url if form_question.image else None,
            'options': options_data
        }
        
        # Si la pregunta es de tipo escala lineal, incluir los valores
        if form_question.question_type == 'linear_scale':
            question_data['min_value'] = form_question.min_value
            question_data['max_value'] = form_question.max_value
            question_data['min_label'] = form_question.min_label
            question_data['max_label'] = form_question.max_label
        
        return JsonResponse({
            'success': True,
            'message': 'Pregunta añadida correctamente al formulario',
            'question': question_data
        })
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"Error en gform_add_from_bank: {str(e)}\n{error_traceback}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

# Añadir una nueva vista para obtener las preguntas en formato JSON
@login_required
def gform_questions_json(request, form_id):
    """Vista para obtener las preguntas de un formulario en formato JSON"""
    gform = get_object_or_404(GForm, id=form_id)
    
    # Verificar que el usuario tenga permisos para ver
    if not gform.can_view(request.user):
        return JsonResponse({'success': False, 'error': 'No tienes permiso para ver este formulario'}, status=403)
    
    # Obtener todas las preguntas del formulario
    questions = GQuestion.objects.filter(form=gform).order_by('position')
    
    # Preparar los datos para la respuesta JSON
    questions_data = []
    for question in questions:
        question_data = {
            'id': question.id,
            'text': question.text,
            'help_text': question.help_text,
            'question_type': question.question_type,
            'question_type_display': question.get_question_type_display(),
            'is_required': question.is_required,
            'allow_attachments': question.allow_attachments,
            'in_question_bank': question.in_question_bank,
            'position': question.position,
            'image': question.image.url if question.image else None,
        }
        
        # Si la pregunta es de tipo selección, incluir las opciones
        if question.question_type in ['multiple_choice', 'checkbox', 'dropdown']:
            options_data = []
            for option in question.options.all():
                options_data.append({
                    'id': option.id,
                    'text': option.text,
                    'position': option.position
                })
            question_data['options'] = options_data
        
        # Si la pregunta es de tipo escala lineal, incluir los valores
        if question.question_type == 'linear_scale':
            question_data['min_value'] = question.min_value
            question_data['max_value'] = question.max_value
            question_data['min_label'] = question.min_label
            question_data['max_label'] = question.max_label
        
        questions_data.append(question_data)
    
    return JsonResponse({
        'success': True,
        'questions': questions_data
    })

# Añadir una nueva vista para limpiar duplicados manualmente
@login_required
def gform_clean_question_bank(request):
    """Vista para limpiar duplicados en el banco de preguntas"""
    try:
        # Obtener todas las preguntas del banco del usuario
        all_questions = BankQuestion.objects.filter(created_by=request.user)
        
        # Usar un diccionario para rastrear preguntas por contenido
        seen_content = {}
        duplicates_to_delete = []
        kept_count = 0
        
        for question in all_questions:
            # Crear una clave única basada en el contenido
            content_key = f"{question.text}|{question.question_type}"
            
            if content_key in seen_content:
                # Este es un duplicado por contenido, marcarlo para eliminación
                duplicates_to_delete.append(question.id)
                print(f"Duplicado detectado: {question.id} (original: {seen_content[content_key].id})")
            else:
                seen_content[content_key] = question
                kept_count += 1
        
        # Eliminar los duplicados detectados
        if duplicates_to_delete:
            delete_count = len(duplicates_to_delete)
            BankQuestion.objects.filter(id__in=duplicates_to_delete).delete()
            print(f"Eliminados {delete_count} duplicados. Mantenidas {kept_count} preguntas únicas.")
        else:
            delete_count = 0
            print("No se encontraron duplicados.")
        
        return JsonResponse({
            'success': True,
            'message': f'Se eliminaron {delete_count} preguntas duplicadas. Quedan {kept_count} preguntas únicas.',
            'duplicates_removed': delete_count,
            'questions_kept': kept_count
        })
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"Error en gform_clean_question_bank: {str(e)}\n{error_traceback}")
        return JsonResponse({
            'success': False,
            'error': str(e),
            'traceback': error_traceback
        }, status=500)

# Nuevas vistas para gestionar permisos de formularios
@login_required
@require_POST
def gform_add_permission(request, form_id):
    """Vista para añadir un permiso a un usuario para un formulario"""
    gform = get_object_or_404(GForm, id=form_id)
    
    # Verificar que el usuario sea el propietario
    if gform.user != request.user:
        return HttpResponseForbidden("Solo el propietario puede gestionar permisos")
    
    form = FormPermissionForm(request.POST)
    if form.is_valid():
        # Obtener el usuario por correo electrónico
        email = form.cleaned_data['user_email']
        try:
            user = User.objects.get(email=email)
            
            # Verificar que no se esté intentando dar permisos al propietario
            if user == gform.user:
                messages.error(request, "No puedes añadir permisos al propietario del formulario")
                return redirect('gform_share_form', form_id=gform.id)
            
            # Verificar que el permiso sea de editor solo si el usuario tiene cuenta
            permission_type = form.cleaned_data['permission_type']
            if permission_type == 'editor' and not user.is_authenticated:
                messages.error(request, "Solo usuarios con cuenta pueden tener permisos de editor")
                return redirect('gform_share_form', form_id=gform.id)
            
            # Crear o actualizar el permiso
            permission, created = FormPermission.objects.update_or_create(
                form=gform,
                user=user,
                defaults={'permission_type': permission_type}
            )
            
            if created:
                messages.success(request, f"Permiso añadido para {user.username}")
            else:
                messages.success(request, f"Permiso actualizado para {user.username}")
            
        except User.DoesNotExist:
            messages.error(request, f"No existe un usuario con el correo electrónico {email}")
        
        return redirect('gform_share_form', form_id=gform.id)
    
    # Si hay errores en el formulario
    for field, errors in form.errors.items():
        for error in errors:
            messages.error(request, f"Error en {field}: {error}")
    
    return redirect('gform_share_form', form_id=gform.id)

@login_required
@require_POST
def gform_remove_permission(request, permission_id):
    """Vista para eliminar un permiso de un formulario"""
    permission = get_object_or_404(FormPermission, id=permission_id)
    gform = permission.form
    
    # Verificar que el usuario sea el propietario
    if gform.user != request.user:
        return HttpResponseForbidden("Solo el propietario puede gestionar permisos")
    
    user_name = permission.user.username
    permission.delete()
    
    messages.success(request, f"Permiso eliminado para {user_name}")
    return redirect('gform_share_form', form_id=gform.id)

@login_required
@require_POST
def gform_create_share_link(request, form_id):
    """Vista para crear un enlace de compartir para un formulario"""
    gform = get_object_or_404(GForm, id=form_id)
    
    # Verificar que el usuario sea el propietario
    if gform.user != request.user:
        return HttpResponseForbidden("Solo el propietario puede crear enlaces de compartir")
    
    form = FormShareLinkForm(request.POST)
    if form.is_valid():
        # Crear el enlace de compartir
        share_link = FormShareLink(
            form=gform,
            permission_type=form.cleaned_data['permission_type']
        )
        
        # Configurar la fecha de expiración si se seleccionó
        expires_in = form.cleaned_data['expires_in']
        if expires_in != 'never':
            days = int(expires_in.replace('d', ''))
            share_link.expires_at = timezone.now() + timedelta(days=days)
        
        share_link.save()
        
        messages.success(request, "Enlace de compartir creado correctamente")
        return redirect('gform_share_form', form_id=gform.id)
    
    # Si hay errores en el formulario
    for field, errors in form.errors.items():
        for error in errors:
            messages.error(request, f"Error en {field}: {error}")
    
    return redirect('gform_share_form', form_id=gform.id)

@login_required
@require_POST
def gform_delete_share_link(request, link_id):
    """Vista para eliminar un enlace de compartir"""
    share_link = get_object_or_404(FormShareLink, id=link_id)
    gform = share_link.form
    
    # Verificar que el usuario sea el propietario
    if gform.user != request.user:
        return HttpResponseForbidden("Solo el propietario puede eliminar enlaces de compartir")
    
    share_link.delete()
    
    messages.success(request, "Enlace de compartir eliminado correctamente")
    return redirect('gform_share_form', form_id=gform.id)

# Vamos a revisar y corregir la función que maneja los enlaces compartidos
def gform_shared_link(request, token):
    """Vista para acceder a un formulario mediante un enlace compartido"""
    share_link = get_object_or_404(FormShareLink, token=token)
    
    # Verificar que el enlace sea válido
    if not share_link.is_valid():
        return render(request, 'forms_google/shared_link_expired.html')
    
    gform = share_link.form
    
    # Si el usuario está autenticado, añadir el permiso correspondiente
    if request.user.is_authenticated:
        # Asignar el permiso exactamente como está definido en el enlace compartido
        # Corregir: Asegurarse de que el permiso se guarda correctamente
        permission, created = FormPermission.objects.update_or_create(
            form=gform,
            user=request.user,
            defaults={'permission_type': share_link.permission_type}
        )
        
        # Añadir log para depuración
        print(f"Permiso asignado: {permission.permission_type} para usuario {request.user.username} en formulario {gform.id}")
        
        permission_display = dict(FormPermission.PERMISSION_CHOICES).get(share_link.permission_type, share_link.permission_type)
        messages.success(request, f"Ahora tienes permisos de {permission_display.lower()} para el formulario '{gform.title}'")
        
        # Redirigir según el tipo de permiso
        if share_link.permission_type == 'editor':
            return redirect('gform_edit', form_id=gform.id)
        elif share_link.permission_type == 'responder':
            return redirect('gform_respond', form_id=gform.id)
        else:  # viewer
            return redirect('gform_view', form_id=gform.id)
    else:
        # Si el permiso es de editor, redirigir al login
        if share_link.permission_type == 'editor':
            # Guardar la URL de redirección en la sesión
            request.session['next'] = reverse('gform_edit', kwargs={'form_id': gform.id})
            messages.info(request, "Debes iniciar sesión para editar este formulario")
            return redirect('login')
        
        # Para responder o ver, redirigir directamente
        if share_link.permission_type == 'responder':
            return redirect('gform_respond', form_id=gform.id)
        else:  # viewer
            return redirect('gform_view', form_id=gform.id)

@login_required
@require_POST
def gform_add_shared_link(request):
    """Vista para agregar un formulario compartido mediante un enlace"""
    shared_link = request.POST.get('shared_link', '').strip()
    
    if not shared_link:
        messages.error(request, "Debes proporcionar un enlace válido")
        return redirect('gform_list')
    
    # Extraer el token del enlace
    import re
    token_match = re.search(r'/shared/([a-f0-9-]+)', shared_link)
    
    if not token_match:
        messages.error(request, "El enlace proporcionado no es válido")
        return redirect('gform_list')
    
    token = token_match.group(1)
    
    try:
        # Buscar el enlace compartido
        share_link = get_object_or_404(FormShareLink, token=uuid.UUID(token))
        
        # Verificar que el enlace sea válido
        if not share_link.is_valid():
            messages.error(request, "El enlace ha expirado")
            return redirect('gform_list')
        
        gform = share_link.form
        
        # Verificar que no sea el propietario
        if gform.user == request.user:
            messages.warning(request, "No puedes agregar tu propio formulario como compartido")
            return redirect('gform_list')
        
        # Verificar si ya tiene permisos
        existing_permission = FormPermission.objects.filter(form=gform, user=request.user).first()
        
        if existing_permission:
            # Si ya tiene permisos, actualizar solo si el nuevo es de mayor nivel
            permission_levels = {'viewer': 1, 'responder': 2, 'editor': 3}
            current_level = permission_levels.get(existing_permission.permission_type, 0)
            new_level = permission_levels.get(share_link.permission_type, 0)
            
            if new_level > current_level:
                existing_permission.permission_type = share_link.permission_type
                existing_permission.save()
                messages.success(request, f"Tus permisos para '{gform.title}' han sido actualizados a {share_link.get_permission_type_display()}")
            else:
                messages.info(request, f"Ya tienes permisos para '{gform.title}'")
        else:
            # Si el permiso es de editor, verificar que el usuario tenga cuenta
            if share_link.permission_type == 'editor' and not request.user.is_authenticated:
                messages.error(request, "Solo usuarios con cuenta pueden tener permisos de editor")
                return redirect('gform_list')
            
            # Crear el permiso
            FormPermission.objects.create(
                form=gform,
                user=request.user,
                permission_type=share_link.permission_type
            )
            messages.success(request, f"Formulario '{gform.title}' agregado a tus formularios compartidos")
        
        return redirect('gform_list')
    
    except (ValueError, FormShareLink.DoesNotExist):
        messages.error(request, "El enlace proporcionado no es válido")
        return redirect('gform_list')

@login_required
def gform_remove_shared(request, form_id):
    """Vista para eliminar un formulario de la lista de compartidos"""
    gform = get_object_or_404(GForm, id=form_id)
    
    # Verificar que no sea el propietario
    if gform.user == request.user:
        messages.warning(request, "No puedes eliminar tu propio formulario de compartidos")
        return redirect('gform_list')
    
    # Eliminar el permiso
    FormPermission.objects.filter(form=gform, user=request.user).delete()
    
    messages.success(request, f"Formulario '{gform.title}' eliminado de tus formularios compartidos")
    return redirect('gform_list')

def change_language(request):
    """
    Vista para cambiar el idioma de la aplicación
    """
    from django.http import HttpResponseRedirect
    import logging
    
    # Configurar logging
    logger = logging.getLogger(__name__)
    
    language = request.POST.get('language', 'es')
    next_url = request.POST.get('next', '/')
    
    # Guardar el idioma en la sesión
    request.session['language'] = language
    
    # Forzar que la sesión se guarde inmediatamente
    request.session.modified = True
    
    # Imprimir para depuración
    logger.debug(f"change_language: Cambiando idioma a {language}, redirigiendo a {next_url}")
    print(f"change_language: Cambiando idioma a {language}, redirigiendo a {next_url}")
    print(f"Session: {request.session.items()}")
    
    # Establecer también en el contexto global
    from django.utils import translation
    translation.activate(language)
    
    # Añadir el idioma como cookie
    response = HttpResponseRedirect(next_url)
    response.set_cookie('django_language', language, max_age=60*60*24*365)
    
    return response

# Modificar la función can_respond para que solo los editores y respondedores puedan responder
def can_respond(self, user):
    """Verifica si el usuario puede responder el formulario"""
    permission = self.get_user_permission(user)
    return permission in ['owner', 'editor', 'responder'] or self.is_published